#!/usr/bin/env python3
"""
seed_students.py — xlsx → Supabase students 테이블 초기 적재

사전 준비:
  pip install openpyxl requests

Supabase SQL (실행 순서):
  1) CREATE TABLE students (
       id          uuid PRIMARY KEY DEFAULT gen_random_uuid(),
       class_num   int  NOT NULL,
       student_num int  NOT NULL,
       name        text NOT NULL,
       study_room  text NOT NULL,
       schedule    jsonb NOT NULL DEFAULT '{}',
       created_at  timestamptz DEFAULT now()
     );

사용법:
  python scripts/seed_students.py [xlsx경로]

xlsx 기본값: ./data/2026. 청·백운반 출석부.xlsx
"""

import sys
import json
import requests

try:
    import openpyxl
except ImportError:
    print("ERROR: openpyxl이 설치되지 않았습니다.\n  pip install openpyxl requests")
    sys.exit(1)

# ─── 설정 (js/config.js와 동일한 값 입력) ─────────────────
SUPABASE_URL = 'https://crlclfzueyurkoqcrojc.supabase.co'
SUPABASE_KEY = 'eyJhbGciOiJIUzI1NiIsInR5cCI6IkpXVCJ9.eyJpc3MiOiJzdXBhYmFzZSIsInJlZiI6ImNybGNsZnp1ZXl1cmtvcWNyb2pjIiwicm9sZSI6ImFub24iLCJpYXQiOjE3ODI2Njg0NTIsImV4cCI6MjA5ODI0NDQ1Mn0.0iVRFhqy64H6GIFcCDeQWQDx0Tt9vb5LkwKNghNnYp4'

XLSX_PATH = sys.argv[1] if len(sys.argv) > 1 else "./data/2026. 청·백운반 출석부.xlsx"
# ──────────────────────────────────────────────────────────


def cell_val(v):
    """xlsx 셀 값 → 스케줄 문자열 변환 (None → '-')"""
    if v is None:
        return "-"
    s = str(v).strip()
    return s if s in ("O", "방과후") else "-"


def parse_students(xlsx_path: str) -> list[dict]:
    wb = openpyxl.load_workbook(xlsx_path, data_only=True)
    ws = wb.worksheets[0]  # 시트1

    # 컬럼 레이아웃 (0-indexed):
    # 0=반, 1=번호, 2=이름
    # 3-5  = 월(오자/야자/심자)
    # 6-8  = 화
    # 9-11 = 수
    # 12-14= 목
    # 15-17= 금
    # 18-19= 토(오전/오후)
    # 20   = 자습반

    students = []
    for row in ws.iter_rows(min_row=3, max_row=ws.max_row, max_col=21, values_only=True):
        if row[0] is None or row[2] is None:
            continue

        study_room = str(row[20]).strip() if row[20] else None
        if not study_room:
            continue

        schedule = {
            "mon": [cell_val(row[3]),  cell_val(row[4]),  cell_val(row[5])],
            "tue": [cell_val(row[6]),  cell_val(row[7]),  cell_val(row[8])],
            "wed": [cell_val(row[9]),  cell_val(row[10]), cell_val(row[11])],
            "thu": [cell_val(row[12]), cell_val(row[13]), cell_val(row[14])],
            "fri": [cell_val(row[15]), cell_val(row[16]), cell_val(row[17])],
            "sat": [cell_val(row[18]), cell_val(row[19])],
        }

        students.append({
            "class_num":   int(row[0]),
            "student_num": int(row[1]),
            "name":        str(row[2]).strip(),
            "study_room":  study_room,
            "schedule":    schedule,
        })

    return students


def insert_students(students: list[dict]) -> None:
    url = f"{SUPABASE_URL}/rest/v1/students"
    headers = {
        "apikey":        SUPABASE_KEY,
        "Authorization": f"Bearer {SUPABASE_KEY}",
        "Content-Type":  "application/json",
        "Prefer":        "return=minimal",
    }

    BATCH = 100
    total = 0
    for i in range(0, len(students), BATCH):
        batch = students[i : i + BATCH]
        resp = requests.post(url, headers=headers, data=json.dumps(batch, ensure_ascii=False).encode("utf-8"))
        if resp.status_code not in (200, 201):
            print(f"ERROR {resp.status_code}: {resp.text}")
            sys.exit(1)
        total += len(batch)
        print(f"  Inserted {total}/{len(students)}")

    print(f"\n완료: {len(students)}명 적재")


def dry_run(students: list[dict]) -> None:
    print("\n[DRY RUN] 처음 3명 미리보기:")
    for s in students[:3]:
        print(json.dumps(s, ensure_ascii=False, indent=2))

    study_rooms = {}
    for s in students:
        study_rooms[s["study_room"]] = study_rooms.get(s["study_room"], 0) + 1
    print("\n자습반별 인원:")
    for room, cnt in sorted(study_rooms.items()):
        print(f"  {room}: {cnt}명")

    print("\nSUPABASE_URL이 플레이스홀더입니다.")
    print("scripts/seed_students.py 상단의 SUPABASE_URL / SUPABASE_KEY를 입력하고 다시 실행하세요.")


def main() -> None:
    print(f"Reading: {XLSX_PATH}")
    students = parse_students(XLSX_PATH)
    print(f"Parsed: {len(students)}명")

    if "YOUR_PROJECT" in SUPABASE_URL:
        dry_run(students)
        return

    insert_students(students)


if __name__ == "__main__":
    main()
